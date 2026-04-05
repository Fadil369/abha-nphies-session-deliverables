#!/usr/bin/env python3
"""
BAT-2026-NB-00004295-OT Claims Analysis Spreadsheet Builder
Processes extracted PDF data into a comprehensive multi-sheet Excel workbook.
"""
import json, re, sys
from collections import Counter, defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ─── Config ───
INPUT = '/tmp/bat_4295_extracted.json'
OUTPUT = '/Users/fadil369/abha-nphies-session-deliverables-1/outputs/BAT-2026-NB-00004295-OT_Claims_Analysis.xlsx'

# ─── ART PreAuth Protocol Knowledge Base ───
PREAUTH_PROTOCOL = {
    'BE-1-4': {
        'description': 'Preauthorization is required and was not obtained',
        'protocol_section': 'ART PreAuth – General Requirements',
        'requires_preauth': True,
        'preventable': 'YES',
        'action': 'Submit retroactive pre-authorization request with clinical justification',
        'supporting_docs': 'Medical report, clinical notes, diagnosis confirmation, treatment plan',
        'appeal_strategy': 'Contractual Appeal - Submit retroactive approval request via NPHIES Communication',
        'typical_services': ['Specialist Consultations', 'Chemotherapy', 'Imaging >4 views', 'Medical Devices', 'Dental Specialist'],
    },
    'MN-1-1': {
        'description': 'Service is not clinically justified based on clinical practice guideline',
        'protocol_section': 'ART Clinical Justification – Evidence Requirements',
        'requires_preauth': False,
        'preventable': 'PARTIALLY',
        'action': 'Resubmit with supporting diagnosis and clinical practice guideline reference',
        'supporting_docs': 'Lab results, clinical notes, CPG reference, medical necessity letter',
        'appeal_strategy': 'Resubmit with Supporting Info - Attach clinical evidence per NPHIES guidelines',
        'typical_services': ['Lab Tests', 'Diagnostic workups'],
    },
    'CV-1-3': {
        'description': 'Diagnosis is not covered',
        'protocol_section': 'ART Coverage – Benefit Exclusions',
        'requires_preauth': False,
        'preventable': 'NO',
        'action': 'Review policy exclusion list; if applicable, resubmit under alternate covered diagnosis',
        'supporting_docs': 'Policy benefit schedule, clinical documentation for alternate diagnosis',
        'appeal_strategy': 'New Claim with Prior Linkage if alternate diagnosis applies',
        'typical_services': ['Various – depends on policy exclusions'],
    },
    'BE-1-3': {
        'description': 'Submission not compliant with contractual agreement',
        'protocol_section': 'ART Contractual Compliance',
        'requires_preauth': False,
        'preventable': 'YES',
        'action': 'Correct service code mapping and resubmit with valid codes',
        'supporting_docs': 'Correct NPHIES service code mapping, contract reference',
        'appeal_strategy': 'Resubmit with corrected codes per contractual agreement',
        'typical_services': ['Unknown/Error coded services'],
    },
    'SE-1-6': {
        'description': 'Investigation result is inadequate or missing',
        'protocol_section': 'ART Supporting Evidence – Dental/Diagnostic',
        'requires_preauth': False,
        'preventable': 'YES',
        'action': 'Attach investigation results (X-ray, lab) and resubmit',
        'supporting_docs': 'X-ray images, lab reports, dental charting, clinical photos',
        'appeal_strategy': 'Resubmit with Supporting Info - Attach missing investigation results',
        'typical_services': ['Dental', 'Radiology', 'Diagnostic procedures'],
    },
    'AD-1-4': {
        'description': 'Diagnosis is inconsistent with service/procedure',
        'protocol_section': 'ART Diagnosis-Service Alignment',
        'requires_preauth': False,
        'preventable': 'PARTIALLY',
        'action': 'Review and correct diagnosis-service mapping; resubmit with consistent ICD codes',
        'supporting_docs': 'Updated clinical notes with correct ICD-10 codes, treatment rationale',
        'appeal_strategy': 'Resubmit with corrected diagnosis codes',
        'typical_services': ['Lab Tests', 'Procedures with mismatched diagnoses'],
    },
    'CV-1-9': {
        'description': 'Service requires additional coverage verification',
        'protocol_section': 'ART Coverage Verification',
        'requires_preauth': False,
        'preventable': 'PARTIALLY',
        'action': 'Verify member eligibility and benefit coverage before resubmission',
        'supporting_docs': 'Eligibility verification, benefit schedule confirmation',
        'appeal_strategy': 'Resubmit after eligibility/coverage verification',
        'typical_services': ['Various'],
    },
    'AD-3-7': {
        'description': 'Additional information required for adjudication',
        'protocol_section': 'ART Adjudication Support Documents',
        'requires_preauth': False,
        'preventable': 'YES',
        'action': 'Provide requested additional documentation and resubmit',
        'supporting_docs': 'As specified in rejection detail – typically medical reports, clinical notes',
        'appeal_strategy': 'Resubmit with Supporting Info',
        'typical_services': ['Various'],
    },
    'AD-2-4': {
        'description': 'Duplicate submission detected',
        'protocol_section': 'ART Duplicate Detection',
        'requires_preauth': False,
        'preventable': 'YES',
        'action': 'Verify original claim status; do not resubmit if already processed',
        'supporting_docs': 'Original claim reference, settlement report',
        'appeal_strategy': 'Verify and reconcile – appeal only if original was not settled',
        'typical_services': ['Any'],
    },
}

# ─── Styles ───
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(name='Calibri', bold=True, size=14, color='1F4E79')
SUBTITLE_FONT = Font(name='Calibri', bold=True, size=11, color='4472C4')
DATA_FONT = Font(name='Calibri', size=10)
MONEY_FMT = '#,##0.00'
PCT_FMT = '0.0%'
THIN_BORDER = Border(
    left=Side(style='thin', color='D9E2F3'),
    right=Side(style='thin', color='D9E2F3'),
    top=Side(style='thin', color='D9E2F3'),
    bottom=Side(style='thin', color='D9E2F3'),
)

SEVERITY_COLORS = {
    'YES': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    'PARTIALLY': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
    'NO': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
}

def style_header_row(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def style_data_cell(cell, wrap=False):
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    if wrap:
        cell.alignment = Alignment(wrap_text=True, vertical='top')

def auto_width(ws, min_w=8, max_w=45):
    for col in ws.columns:
        mx = min_w
        for cell in col:
            if cell.value:
                mx = max(mx, min(len(str(cell.value)), max_w))
        ws.column_dimensions[get_column_letter(col[0].column)].width = mx + 2

# ─── Load data ───
with open(INPUT) as f:
    data = json.load(f)
header = data['header']
claims = data['claims']

wb = Workbook()

# ═══════════════════════════════════════════
# Sheet 1: Batch Summary
# ═══════════════════════════════════════════
ws = wb.active
ws.title = 'Batch Summary'
ws.sheet_properties.tabColor = '1F4E79'

ws.merge_cells('A1:F1')
ws['A1'] = 'REJECTED CLAIMS ANALYSIS'
ws['A1'].font = TITLE_FONT

ws.merge_cells('A2:F2')
ws['A2'] = f'Batch: {header.get("batch_no", "N/A")} | Period: {header.get("from_date", "")} to {header.get("to_date", "")} | Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
ws['A2'].font = SUBTITLE_FONT

# Financial summary
fin_headers = ['Metric', 'Amount (SAR)', 'Percentage']
for ci, h in enumerate(fin_headers, 1):
    ws.cell(row=4, column=ci, value=h)
style_header_row(ws, 4, len(fin_headers))

net_claimed = float(header.get('net_claimed', '0').replace(',', ''))
rejected = float(header.get('rejected_amount', '0').replace(',', ''))
approved = float(header.get('approved_amount', '0').replace(',', ''))
vat_approved = float(header.get('vat_approved', '0').replace(',', ''))
vat_rejected = float(header.get('vat_rejected', '0').replace(',', ''))
gross_var = float(header.get('gross_variance', '0').replace(',', ''))
total_claims = float(header.get('total_claims', '0').replace(',', ''))

fin_data = [
    ('Net Claimed', net_claimed, 1.0),
    ('Approved Amount', approved, approved / net_claimed if net_claimed else 0),
    ('Rejected Amount', rejected, rejected / net_claimed if net_claimed else 0),
    ('VAT Approved', vat_approved, None),
    ('VAT Rejected', vat_rejected, None),
    ('Gross Amount Variance', gross_var, None),
    ('Total Claims in Batch', total_claims, None),
    ('Rejected Line Items (extracted)', len(claims), None),
    ('Unique Bundle IDs', len(set(c['bundle_id'] for c in claims)), None),
    ('Unique Members', len(set(c['national_id'] for c in claims)), None),
    ('Overall Rejection Rate', None, rejected / net_claimed if net_claimed else 0),
    ('Approval Rate', None, approved / net_claimed if net_claimed else 0),
]

for ri, (metric, amt, pct) in enumerate(fin_data, 5):
    c1 = ws.cell(row=ri, column=1, value=metric)
    style_data_cell(c1)
    if amt is not None:
        c2 = ws.cell(row=ri, column=2, value=amt)
        c2.number_format = MONEY_FMT
        style_data_cell(c2)
    if pct is not None:
        c3 = ws.cell(row=ri, column=3, value=pct)
        c3.number_format = PCT_FMT
        style_data_cell(c3)

# Rejection code summary
r = len(fin_data) + 7
ws.cell(row=r, column=1, value='REJECTION CODE BREAKDOWN').font = SUBTITLE_FONT
r += 1
code_headers = ['Rejection Code', 'Description', 'Count', '% of Total', 'Preventable', 'Recommended Action']
for ci, h in enumerate(code_headers, 1):
    ws.cell(row=r, column=ci, value=h)
style_header_row(ws, r, len(code_headers))

rej_counts = Counter(c['rejection_code'] for c in claims)
for code, cnt in rej_counts.most_common():
    r += 1
    proto = PREAUTH_PROTOCOL.get(code, {})
    ws.cell(row=r, column=1, value=code)
    ws.cell(row=r, column=2, value=proto.get('description', 'Unknown'))
    ws.cell(row=r, column=3, value=cnt)
    c4 = ws.cell(row=r, column=4, value=cnt / len(claims))
    c4.number_format = PCT_FMT
    prev = proto.get('preventable', 'N/A')
    c5 = ws.cell(row=r, column=5, value=prev)
    if prev in SEVERITY_COLORS:
        c5.fill = SEVERITY_COLORS[prev]
    ws.cell(row=r, column=6, value=proto.get('action', ''))
    for ci in range(1, 7):
        style_data_cell(ws.cell(row=r, column=ci), wrap=(ci in [2, 6]))

# Category summary
r += 3
ws.cell(row=r, column=1, value='SERVICE CATEGORY BREAKDOWN').font = SUBTITLE_FONT
r += 1
cat_headers = ['Category', 'Count', '% of Total']
for ci, h in enumerate(cat_headers, 1):
    ws.cell(row=r, column=ci, value=h)
style_header_row(ws, r, len(cat_headers))

cat_counts = Counter(c['category'] for c in claims)
for cat, cnt in cat_counts.most_common():
    r += 1
    ws.cell(row=r, column=1, value=cat)
    ws.cell(row=r, column=2, value=cnt)
    c3 = ws.cell(row=r, column=3, value=cnt / len(claims))
    c3.number_format = PCT_FMT
    for ci in range(1, 4):
        style_data_cell(ws.cell(row=r, column=ci))

auto_width(ws)

# ═══════════════════════════════════════════
# Sheet 2: Claims Data (full normalized extract)
# ═══════════════════════════════════════════
ws2 = wb.create_sheet('Claims Data')
ws2.sheet_properties.tabColor = '4472C4'

headers2 = ['Bundle ID', 'Member Name', 'National ID', 'Service Date', 'Service Code',
            'Service Name', 'Rejection Code', 'Rejection Description', 'Category',
            'Requires PreAuth', 'Preventable', 'Recommended Action']
for ci, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=ci, value=h)
style_header_row(ws2, 1, len(headers2))

for ri, c in enumerate(claims, 2):
    proto = PREAUTH_PROTOCOL.get(c['rejection_code'], {})
    vals = [
        c['bundle_id'], c['member_name'], c['national_id'], c['service_date'],
        c['service_code'], c['service_name'], c['rejection_code'], c['rejection_description'],
        c['category'],
        'YES' if proto.get('requires_preauth') else 'NO',
        proto.get('preventable', 'N/A'),
        proto.get('action', ''),
    ]
    for ci, v in enumerate(vals, 1):
        cell = ws2.cell(row=ri, column=ci, value=v)
        style_data_cell(cell, wrap=(ci in [6, 8, 12]))
        # Color preventable column
        if ci == 11 and v in SEVERITY_COLORS:
            cell.fill = SEVERITY_COLORS[v]

ws2.auto_filter.ref = f'A1:L{len(claims)+1}'
ws2.freeze_panes = 'A2'
auto_width(ws2)

# ═══════════════════════════════════════════
# Sheet 3: Appeal Tracker
# ═══════════════════════════════════════════
ws3 = wb.create_sheet('Appeal Tracker')
ws3.sheet_properties.tabColor = 'ED7D31'

headers3 = ['Bundle ID', 'Member Name', 'National ID', 'Rejection Code', 'Category',
            'Appeal Strategy', 'Supporting Docs Required', 'Protocol Section',
            'Status', 'Notes']
for ci, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=ci, value=h)
style_header_row(ws3, 1, len(headers3))

# One row per unique bundle+rejection_code combo for appeal tracking
seen = set()
arow = 2
for c in claims:
    key = (c['bundle_id'], c['rejection_code'])
    if key in seen:
        continue
    seen.add(key)
    proto = PREAUTH_PROTOCOL.get(c['rejection_code'], {})
    vals = [
        c['bundle_id'], c['member_name'], c['national_id'], c['rejection_code'], c['category'],
        proto.get('appeal_strategy', 'Review and assess'),
        proto.get('supporting_docs', 'Clinical documentation'),
        proto.get('protocol_section', 'N/A'),
        'PENDING',
        '',
    ]
    for ci, v in enumerate(vals, 1):
        cell = ws3.cell(row=arow, column=ci, value=v)
        style_data_cell(cell, wrap=(ci in [6, 7, 8]))
    arow += 1

ws3.auto_filter.ref = f'A1:J{arow-1}'
ws3.freeze_panes = 'A2'
auto_width(ws3)

# ═══════════════════════════════════════════
# Sheet 4: Priority Actions
# ═══════════════════════════════════════════
ws4 = wb.create_sheet('Priority Actions')
ws4.sheet_properties.tabColor = '00B050'

ws4.merge_cells('A1:G1')
ws4['A1'] = 'PRIORITY ACTION ITEMS - IMMEDIATE RESPONSE PLAN'
ws4['A1'].font = TITLE_FONT

headers4 = ['Priority', 'Rejection Code', 'Count', '% Impact', 'Preventable',
            'Immediate Action', 'Expected Recovery (SAR)', 'Deadline']
for ci, h in enumerate(headers4, 1):
    ws4.cell(row=3, column=ci, value=h)
style_header_row(ws4, 3, len(headers4))

# Estimate value per rejected item
avg_rejection_value = rejected / len(claims) if claims else 0

prow = 4
for priority, (code, cnt) in enumerate(rej_counts.most_common(), 1):
    proto = PREAUTH_PROTOCOL.get(code, {})
    est_recovery = avg_rejection_value * cnt * (0.7 if proto.get('preventable') == 'YES' else 0.3 if proto.get('preventable') == 'PARTIALLY' else 0.05)
    
    ws4.cell(row=prow, column=1, value=priority)
    ws4.cell(row=prow, column=2, value=code)
    ws4.cell(row=prow, column=3, value=cnt)
    c4 = ws4.cell(row=prow, column=4, value=cnt / len(claims))
    c4.number_format = PCT_FMT
    prev = proto.get('preventable', 'N/A')
    c5 = ws4.cell(row=prow, column=5, value=prev)
    if prev in SEVERITY_COLORS:
        c5.fill = SEVERITY_COLORS[prev]
    ws4.cell(row=prow, column=6, value=proto.get('action', 'Investigate'))
    c7 = ws4.cell(row=prow, column=7, value=round(est_recovery, 2))
    c7.number_format = MONEY_FMT
    ws4.cell(row=prow, column=8, value='Within 30 days')
    for ci in range(1, 9):
        style_data_cell(ws4.cell(row=prow, column=ci), wrap=(ci == 6))
    prow += 1

# Total expected recovery
prow += 1
ws4.cell(row=prow, column=6, value='TOTAL EXPECTED RECOVERY').font = Font(bold=True)
total_recovery = sum(
    avg_rejection_value * cnt * (0.7 if PREAUTH_PROTOCOL.get(code, {}).get('preventable') == 'YES' else 0.3 if PREAUTH_PROTOCOL.get(code, {}).get('preventable') == 'PARTIALLY' else 0.05)
    for code, cnt in rej_counts.items()
)
c = ws4.cell(row=prow, column=7, value=round(total_recovery, 2))
c.number_format = MONEY_FMT
c.font = Font(bold=True, color='00B050')

auto_width(ws4)

# ═══════════════════════════════════════════
# Sheet 5: Member Analysis
# ═══════════════════════════════════════════
ws5 = wb.create_sheet('Member Analysis')
ws5.sheet_properties.tabColor = '7030A0'

headers5 = ['National ID', 'Member Name', 'Total Rejected Items', 'Unique Bundles',
            'Primary Rejection Code', 'Categories Affected', 'Risk Level']
for ci, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=ci, value=h)
style_header_row(ws5, 1, len(headers5))

member_data = defaultdict(lambda: {'names': set(), 'bundles': set(), 'codes': [], 'cats': set()})
for c in claims:
    nid = c['national_id']
    member_data[nid]['names'].add(c['member_name'])
    member_data[nid]['bundles'].add(c['bundle_id'])
    member_data[nid]['codes'].append(c['rejection_code'])
    member_data[nid]['cats'].add(c['category'])

mrow = 2
for nid, md in sorted(member_data.items(), key=lambda x: len(x[1]['codes']), reverse=True):
    code_counter = Counter(md['codes'])
    primary = code_counter.most_common(1)[0][0]
    total = len(md['codes'])
    risk = 'HIGH' if total >= 10 else 'MEDIUM' if total >= 5 else 'LOW'
    
    ws5.cell(row=mrow, column=1, value=nid)
    ws5.cell(row=mrow, column=2, value=', '.join(md['names']))
    ws5.cell(row=mrow, column=3, value=total)
    ws5.cell(row=mrow, column=4, value=len(md['bundles']))
    ws5.cell(row=mrow, column=5, value=primary)
    ws5.cell(row=mrow, column=6, value=', '.join(sorted(md['cats'])))
    risk_cell = ws5.cell(row=mrow, column=7, value=risk)
    if risk == 'HIGH':
        risk_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    elif risk == 'MEDIUM':
        risk_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    else:
        risk_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    for ci in range(1, 8):
        style_data_cell(ws5.cell(row=mrow, column=ci), wrap=(ci in [2, 6]))
    mrow += 1

ws5.auto_filter.ref = f'A1:G{mrow-1}'
ws5.freeze_panes = 'A2'
auto_width(ws5)

# ═══════════════════════════════════════════
# Sheet 6: ART Protocol Reference
# ═══════════════════════════════════════════
ws6 = wb.create_sheet('ART Protocol Reference')
ws6.sheet_properties.tabColor = 'BF8F00'

headers6 = ['Rejection Code', 'Description', 'Protocol Section', 'Requires PreAuth',
            'Preventable', 'Supporting Docs Required', 'Appeal Strategy', 'Typical Services']
for ci, h in enumerate(headers6, 1):
    ws6.cell(row=1, column=ci, value=h)
style_header_row(ws6, 1, len(headers6))

prow = 2
for code, proto in PREAUTH_PROTOCOL.items():
    vals = [
        code, proto['description'], proto['protocol_section'],
        'YES' if proto['requires_preauth'] else 'NO',
        proto['preventable'], proto['supporting_docs'],
        proto['appeal_strategy'],
        ', '.join(proto.get('typical_services', [])),
    ]
    for ci, v in enumerate(vals, 1):
        cell = ws6.cell(row=prow, column=ci, value=v)
        style_data_cell(cell, wrap=(ci in [2, 5, 6, 7, 8]))
        if ci == 5 and v in SEVERITY_COLORS:
            cell.fill = SEVERITY_COLORS[v]
    prow += 1

auto_width(ws6)

# ═══════════════════════════════════════════
# Sheet 7: Trend Analysis
# ═══════════════════════════════════════════
ws7 = wb.create_sheet('Trend & Patterns')
ws7.sheet_properties.tabColor = 'FF0000'

ws7.merge_cells('A1:F1')
ws7['A1'] = 'TREND & PATTERN DISCOVERY'
ws7['A1'].font = TITLE_FONT

# Cross-tab: rejection code by category
r = 3
ws7.cell(row=r, column=1, value='REJECTION CODE × CATEGORY MATRIX').font = SUBTITLE_FONT
r += 1
cats = sorted(set(c['category'] for c in claims))
ws7.cell(row=r, column=1, value='Code \\ Category')
for ci, cat in enumerate(cats, 2):
    ws7.cell(row=r, column=ci, value=cat)
ws7.cell(row=r, column=len(cats)+2, value='Total')
style_header_row(ws7, r, len(cats)+2)

cross = defaultdict(lambda: defaultdict(int))
for c in claims:
    cross[c['rejection_code']][c['category']] += 1

for code in sorted(cross.keys()):
    r += 1
    ws7.cell(row=r, column=1, value=code)
    total = 0
    for ci, cat in enumerate(cats, 2):
        v = cross[code].get(cat, 0)
        if v:
            ws7.cell(row=r, column=ci, value=v)
        total += v
    ws7.cell(row=r, column=len(cats)+2, value=total)
    for ci in range(1, len(cats)+3):
        style_data_cell(ws7.cell(row=r, column=ci))

# Date distribution
r += 3
ws7.cell(row=r, column=1, value='REJECTION DISTRIBUTION BY SERVICE DATE').font = SUBTITLE_FONT
r += 1
ws7.cell(row=r, column=1, value='Service Date')
ws7.cell(row=r, column=2, value='Rejections')
ws7.cell(row=r, column=3, value='Unique Members')
style_header_row(ws7, r, 3)

date_data = defaultdict(lambda: {'count': 0, 'members': set()})
for c in claims:
    d = c['service_date']
    date_data[d]['count'] += 1
    date_data[d]['members'].add(c['national_id'])

for d in sorted(date_data.keys()):
    r += 1
    ws7.cell(row=r, column=1, value=d)
    ws7.cell(row=r, column=2, value=date_data[d]['count'])
    ws7.cell(row=r, column=3, value=len(date_data[d]['members']))
    for ci in range(1, 4):
        style_data_cell(ws7.cell(row=r, column=ci))

# Financial exposure
r += 3
ws7.cell(row=r, column=1, value='FINANCIAL EXPOSURE ANALYSIS').font = SUBTITLE_FONT
r += 1
exposure_headers = ['Metric', 'Value (SAR)']
for ci, h in enumerate(exposure_headers, 1):
    ws7.cell(row=r, column=ci, value=h)
style_header_row(ws7, r, 2)

exposure = [
    ('Net Claimed', net_claimed),
    ('Approved Amount', approved),
    ('Rejected Amount', rejected),
    ('Financial Gap (Net − Approved)', net_claimed - approved),
    ('Rejection Rate', None),
    ('Approval Rate', None),
    ('Average Rejection per Line Item', rejected / len(claims) if claims else 0),
    ('Estimated Recoverable (via appeals)', round(total_recovery, 2)),
    ('Net Exposure After Recovery', round(rejected - total_recovery, 2)),
]

for metric, val in exposure:
    r += 1
    ws7.cell(row=r, column=1, value=metric)
    if val is not None:
        c = ws7.cell(row=r, column=2, value=val)
        c.number_format = MONEY_FMT
    elif 'Rate' in metric:
        rate = rejected/net_claimed if 'Rejection' in metric else approved/net_claimed
        c = ws7.cell(row=r, column=2, value=rate)
        c.number_format = PCT_FMT
    for ci in range(1, 3):
        style_data_cell(ws7.cell(row=r, column=ci))

auto_width(ws7)

# ─── Save ───
wb.save(OUTPUT)
print(f'Workbook saved: {OUTPUT}')
print(f'  Sheets: {wb.sheetnames}')
print(f'  Claims Data rows: {len(claims)}')
print(f'  Appeal Tracker entries: {len(seen)}')
print(f'  Members: {len(member_data)}')
print(f'  Estimated recovery: SAR {total_recovery:,.2f}')
