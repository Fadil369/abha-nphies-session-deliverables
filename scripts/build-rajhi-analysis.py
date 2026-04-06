#!/usr/bin/env python3
"""
BAT-2026-NB-00004295-OT  Al-Rajhi Riyadh  Claims Analysis Builder
===================================================================
Full pipeline:
  1. Load pre-extracted JSON  →  structured normalization
  2. Enrich with ART PreAuth Protocol knowledge base
  3. Cross-reference Auth & Claims spreadsheets from the ZIP export
  4. Build 9-sheet Excel workbook with:
       Batch Summary | Claims Data | Appeal Tracker | Priority Actions
       Member Analysis | Provider Comparison | ART Reference
       Trend & Patterns | Submission Queue
  5. Write appeal-prep CSV files ready for NPHIES FHIR bundle generation
"""
import json, re, sys, os
from collections import Counter, defaultdict
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Paths ───────────────────────────────────────────────────────────────────
BASE     = "/Users/fadil369/abha-nphies-session-deliverables-1"
DL       = "/Users/fadil369/Downloads/rajhi-riyad-feb"
JSON_IN  = "/tmp/bat_4295_extracted.json"
AUTH_XLS = f"{DL}/Auth_c34597bf-d184-4ace-a89f-4683dc88b64e_part1_2026-01-25_2026-03-10.xlsx"
CLAIM_XLS= f"{DL}/Claim_9f332ffe-7145-49ee-aa78-5dc56d4cdfc1_part1_2026-02-07_2026-03-10.xlsx"
OUT_XLS  = f"{BASE}/outputs/BAT-2026-NB-00004295-OT_Claims_Analysis.xlsx"
PREP_DIR = f"{BASE}/outputs/rajhi-appeal-prep"
os.makedirs(PREP_DIR, exist_ok=True)


# ─── ART PreAuth Protocol Knowledge Base ─────────────────────────────────────
ART = {
    'BE-1-4': {
        'desc': 'Preauthorization is required and was not obtained',
        'section': 'ART PreAuth – General Requirements',
        'preauth': True, 'preventable': 'YES',
        'action': 'Submit retroactive pre-authorization request with clinical justification',
        'docs': 'Medical report, clinical notes, diagnosis confirmation, treatment plan',
        'strategy': 'Contractual Appeal – retroactive approval via NPHIES CommunicationRequest',
        'services': ['Specialist Consultations', 'Chemotherapy', 'Imaging >4 views',
                     'Medical Devices', 'Dental Specialist', 'Nebulizer/Inhalation'],
        'recovery_rate': 0.70,
    },
    'MN-1-1': {
        'desc': 'Service not clinically justified – no supporting diagnosis',
        'section': 'ART Clinical Justification – Evidence Requirements',
        'preauth': False, 'preventable': 'PARTIALLY',
        'action': 'Resubmit with supporting diagnosis + CPG reference',
        'docs': 'Lab results, clinical notes, CPG reference, medical necessity letter',
        'strategy': 'Resubmit with Supporting Info – attach clinical evidence',
        'services': ['Lab Tests', 'Diagnostic workups'],
        'recovery_rate': 0.40,
    },
    'CV-1-3': {
        'desc': 'Diagnosis is not covered under policy',
        'section': 'ART Coverage – Benefit Exclusions',
        'preauth': False, 'preventable': 'NO',
        'action': 'Review policy exclusion; resubmit under alternate covered diagnosis if applicable',
        'docs': 'Policy benefit schedule, clinical documentation for alternate diagnosis',
        'strategy': 'New Claim with Prior Linkage if alternate diagnosis applies',
        'services': ['Various – policy-dependent'],
        'recovery_rate': 0.10,
    },
    'BE-1-3': {
        'desc': 'Submission not compliant with contractual agreement',
        'section': 'ART Contractual Compliance',
        'preauth': False, 'preventable': 'YES',
        'action': 'Correct service code mapping and resubmit with valid codes',
        'docs': 'Correct NPHIES service code mapping, contract reference',
        'strategy': 'Resubmit with corrected codes per contractual agreement',
        'services': ['Unknown/Error coded services'],
        'recovery_rate': 0.65,
    },
    'SE-1-6': {
        'desc': 'Investigation result is inadequate or missing',
        'section': 'ART Supporting Evidence – Dental/Diagnostic',
        'preauth': False, 'preventable': 'YES',
        'action': 'Attach investigation results (X-ray, lab) and resubmit',
        'docs': 'X-ray images, lab reports, dental charting, clinical photos',
        'strategy': 'Resubmit with Supporting Info – attach missing investigation results',
        'services': ['Dental', 'Radiology', 'Diagnostic procedures'],
        'recovery_rate': 0.60,
    },
    'AD-1-4': {
        'desc': 'Diagnosis code inconsistent with service/procedure',
        'section': 'ART Diagnosis-Service Alignment',
        'preauth': False, 'preventable': 'PARTIALLY',
        'action': 'Correct ICD-10 code mapping; resubmit with consistent diagnosis',
        'docs': 'Updated clinical notes with correct ICD-10 codes, treatment rationale',
        'strategy': 'Resubmit with corrected diagnosis codes',
        'services': ['Lab Tests', 'Procedures with mismatched diagnoses'],
        'recovery_rate': 0.45,
    },
    'CV-1-9': {
        'desc': 'Service requires additional coverage verification',
        'section': 'ART Coverage Verification',
        'preauth': False, 'preventable': 'PARTIALLY',
        'action': 'Verify member eligibility and benefit coverage before resubmission',
        'docs': 'Eligibility verification, benefit schedule confirmation',
        'strategy': 'Resubmit after eligibility/coverage verification',
        'services': ['Various'],
        'recovery_rate': 0.35,
    },
    'AD-3-7': {
        'desc': 'Additional information required for adjudication',
        'section': 'ART Adjudication Support Documents',
        'preauth': False, 'preventable': 'YES',
        'action': 'Provide requested additional documentation and resubmit',
        'docs': 'As specified in rejection – typically medical reports, clinical notes',
        'strategy': 'Resubmit with Supporting Info',
        'services': ['Various'],
        'recovery_rate': 0.55,
    },
    'AD-2-4': {
        'desc': 'Duplicate submission detected',
        'section': 'ART Duplicate Detection',
        'preauth': False, 'preventable': 'YES',
        'action': 'Verify original claim status; do not resubmit if already processed',
        'docs': 'Original claim reference, settlement report',
        'strategy': 'Verify and reconcile – appeal only if original was not settled',
        'services': ['Any'],
        'recovery_rate': 0.20,
    },
    'Override Reason': {
        'desc': 'Manual override reason applied',
        'section': 'N/A',
        'preauth': False, 'preventable': 'PARTIALLY',
        'action': 'Review override context and resubmit with clarification',
        'docs': 'Override justification, clinical notes',
        'strategy': 'Manual Review',
        'services': ['Various'],
        'recovery_rate': 0.30,
    },
}


# ─── Styles ──────────────────────────────────────────────────────────────────
HDR_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HDR_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
TITLE_FONT = Font(name='Calibri', bold=True, size=13, color='1F4E79')
SUB_FONT  = Font(name='Calibri', bold=True, size=11, color='2F75B6')
DATA_FONT = Font(name='Calibri', size=10)
MONEY_FMT, PCT_FMT = '#,##0.00', '0.0%'
THIN = Border(
    left=Side(style='thin',color='D9E2F3'), right=Side(style='thin',color='D9E2F3'),
    top=Side(style='thin',color='D9E2F3'),  bottom=Side(style='thin',color='D9E2F3'),
)
PREV_COLOR = {
    'YES':       PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    'PARTIALLY': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
    'NO':        PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
}
ALT_FILL = PatternFill(start_color='EBF3FB', end_color='EBF3FB', fill_type='solid')


def hdr(ws, row, ncols, fill=HDR_FILL):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT; cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN

def cell_style(cell, wrap=False, money=False, pct=False, alt=False):
    cell.font = DATA_FONT; cell.border = THIN
    if wrap: cell.alignment = Alignment(wrap_text=True, vertical='top')
    if money: cell.number_format = MONEY_FMT
    if pct:   cell.number_format = PCT_FMT
    if alt:   cell.fill = ALT_FILL

def autowidth(ws, mn=8, mx=50):
    for col in ws.columns:
        w = mn
        for cell in col:
            if cell.value:
                w = max(w, min(len(str(cell.value)), mx))
        ws.column_dimensions[get_column_letter(col[0].column)].width = w + 2


# ─── Load extracted JSON ──────────────────────────────────────────────────────
print("Loading extracted JSON...")
with open(JSON_IN) as f:
    data = json.load(f)
header = data['header']
claims = data['claims']

# Filter out the "Override Reason" sentinel if present
claims = [c for c in claims if c.get('bundle_id') and len(c.get('bundle_id','')) > 10]
print(f"  {len(claims)} rejected line items | {len(set(c['bundle_id'] for c in claims))} bundles")

# Enrich each claim with ART protocol data
for c in claims:
    p = ART.get(c['rejection_code'], {})
    c['preauth_required'] = 'YES' if p.get('preauth') else 'NO'
    c['preventable'] = p.get('preventable', 'N/A')
    c['action'] = p.get('action', 'Manual review required')
    c['docs_required'] = p.get('docs', 'Clinical documentation')
    c['appeal_strategy'] = p.get('strategy', 'Manual review')
    c['protocol_section'] = p.get('section', 'N/A')
    c['recovery_rate'] = p.get('recovery_rate', 0.0)


# ─── Load Auth & Claims cross-reference data ─────────────────────────────────
print("Loading Auth data...")
auth_wb = openpyxl.load_workbook(AUTH_XLS, read_only=True)
auth_ws = auth_wb.active
auth_header = list(auth_ws.iter_rows(max_row=1, values_only=True))[0]
auth_rows = list(auth_ws.iter_rows(min_row=2, values_only=True))
auth_map = {}   # bundle_id → row dict
for row in auth_rows:
    d = dict(zip(auth_header, row))
    if d.get('Bundle ID'):
        auth_map[str(d['Bundle ID'])] = d

print(f"  {len(auth_rows)} auth records | {len(auth_map)} unique bundles")

print("Loading Claims data...")
claim_wb = openpyxl.load_workbook(CLAIM_XLS, read_only=True)
claim_ws = claim_wb.active
claim_header = list(claim_ws.iter_rows(max_row=1, values_only=True))[0]
claim_rows = list(claim_ws.iter_rows(min_row=2, values_only=True))
claim_map = {}  # bundle_id → row dict
for row in claim_rows:
    d = dict(zip(claim_header, row))
    if d.get('Bundle ID'):
        claim_map[str(d['Bundle ID'])] = d

print(f"  {len(claim_rows)} claim records | {len(claim_map)} unique bundles")

# Enrich rejected claims with original claim amounts
rejected_bundles = set(c['bundle_id'] for c in claims)
for c in claims:
    cr = claim_map.get(c['bundle_id']) or auth_map.get(c['bundle_id']) or {}
    c['net_amount']       = cr.get('Net Amount') or 0.0
    c['approved_amount']  = cr.get('Approved Amount') or 0.0
    c['claim_status']     = cr.get('Status') or 'Unknown'
    c['policy_number']    = str(cr.get('Policy Number') or '')
    c['patient_id']       = str(cr.get('Patient Identifier') or c.get('national_id',''))
    c['claim_type']       = cr.get('Claim Type') or 'Unknown'
    c['encounter_class']  = cr.get('Encounter Class') or ''
    c['transaction_id']   = str(cr.get('Transaction Identifier') or '')

# Overall portfolio stats (all claims, not just rejected)
all_net      = sum(d.get('Net Amount') or 0 for d in claim_map.values())
all_approved = sum(d.get('Approved Amount') or 0 for d in claim_map.values())
all_statuses = Counter(d.get('Status') for d in claim_map.values() if d.get('Status'))

# Batch header numbers
net_claimed   = float(header['net_claimed'].replace(',',''))
rejected_amt  = float(header['rejected_amount'].replace(',',''))
approved_amt  = float(header['approved_amount'].replace(',',''))
vat_approved  = float(header['vat_approved'].replace(',',''))
vat_rejected  = float(header['vat_rejected'].replace(',',''))
gross_var     = float(header['gross_variance'].replace(',',''))
total_claims  = float(header['total_claims'].replace(',',''))

# Derived stats
rej_per_item  = rejected_amt / len(claims) if claims else 0
rej_counts    = Counter(c['rejection_code'] for c in claims)
cat_counts    = Counter(c['category']       for c in claims)
estimated_recovery = sum(
    rej_counts.get(code, 0) * rej_per_item * ART.get(code, {}).get('recovery_rate', 0)
    for code in rej_counts
)


# ─── Build Workbook ───────────────────────────────────────────────────────────
wb = Workbook()
print("Building Excel workbook...")


# ═══════════════════════════════════════════════════════════
# S1 – BATCH SUMMARY
# ═══════════════════════════════════════════════════════════
ws = wb.active
ws.title = 'Batch Summary'
ws.sheet_properties.tabColor = '1F4E79'

ws.merge_cells('A1:H1')
ws['A1'] = 'AL-RAJHI INSURANCE – REJECTED CLAIMS ANALYSIS'
ws['A1'].font = Font(name='Calibri', bold=True, size=15, color='1F4E79')

ws.merge_cells('A2:H2')
ws['A2'] = (f'Batch: {header["batch_no"]}  |  Period: {header["from_date"]} – {header["to_date"]}'
            f'  |  Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}')
ws['A2'].font = Font(name='Calibri', size=10, color='4472C4', italic=True)

ws.merge_cells('A3:H3')
ws['A3'] = 'Payer: Al-Rajhi Company for Cooperative Insurance  |  Provider: Al-Hayat National Hospital – Riyadh'
ws['A3'].font = Font(name='Calibri', size=10, color='595959')

# Financial KPIs
r = 5
ws.merge_cells(f'A{r}:H{r}')
ws[f'A{r}'] = '▌ BATCH FINANCIAL SUMMARY'
ws[f'A{r}'].font = SUB_FONT

r += 1
cols = ['Metric', 'Amount (SAR)', '% of Net Claimed']
for ci, col in enumerate(cols, 1):
    ws.cell(row=r, column=ci, value=col)
hdr(ws, r, len(cols))

fin_rows = [
    ('Net Claimed',              net_claimed,  1.0),
    ('Approved Amount',          approved_amt, approved_amt/net_claimed),
    ('Rejected Amount',          rejected_amt, rejected_amt/net_claimed),
    ('VAT Approved',             vat_approved, None),
    ('VAT Rejected',             vat_rejected, None),
    ('Gross Variance',           gross_var,    None),
    ('Total Claims in Batch',    total_claims, None),
    ('Rejected Line Items',      len(claims),  None),
    ('Unique Bundles (rejected)',len(rejected_bundles), None),
    ('Unique Members',           len(set(c['national_id'] for c in claims)), None),
    ('Estimated Recovery (appeals)', estimated_recovery, estimated_recovery/rejected_amt if rejected_amt else 0),
    ('Net Exposure After Appeals',   rejected_amt-estimated_recovery, None),
]

for ri, (label, val, pct) in enumerate(fin_rows, r+1):
    alt = ri % 2 == 0
    c1 = ws.cell(row=ri, column=1, value=label); cell_style(c1, alt=alt)
    if isinstance(val, float) and val > 100:
        c2 = ws.cell(row=ri, column=2, value=val); cell_style(c2, money=True, alt=alt)
    elif val is not None:
        c2 = ws.cell(row=ri, column=2, value=val); cell_style(c2, alt=alt)
    if pct is not None:
        c3 = ws.cell(row=ri, column=3, value=pct); cell_style(c3, pct=True, alt=alt)

# Rejection Code Breakdown
r = r + len(fin_rows) + 4
ws.merge_cells(f'A{r}:H{r}')
ws[f'A{r}'] = '▌ REJECTION CODE BREAKDOWN'
ws[f'A{r}'].font = SUB_FONT

r += 1
code_cols = ['Code', 'Description', 'Count', '% of Total', 'Preventable',
             'Requires PreAuth', 'Est. Recovery (SAR)', 'Recommended Action']
for ci, col in enumerate(code_cols, 1):
    ws.cell(row=r, column=ci, value=col)
hdr(ws, r, len(code_cols))

for code, cnt in rej_counts.most_common():
    r += 1
    p = ART.get(code, {})
    est = cnt * rej_per_item * p.get('recovery_rate', 0)
    prev = p.get('preventable','N/A')
    vals = [code, p.get('desc','Unknown'), cnt, cnt/len(claims),
            prev, 'YES' if p.get('preauth') else 'NO', round(est,2),
            p.get('action','Review')]
    for ci, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=ci, value=v)
        cell_style(cell, wrap=(ci in [2,8]), money=(ci==7), pct=(ci==4))
        if ci == 5 and prev in PREV_COLOR:
            cell.fill = PREV_COLOR[prev]

# Category Breakdown
r += 3
ws.merge_cells(f'A{r}:H{r}')
ws[f'A{r}'] = '▌ SERVICE CATEGORY BREAKDOWN'
ws[f'A{r}'].font = SUB_FONT

r += 1
for ci, col in enumerate(['Category','Count','% of Total'], 1):
    ws.cell(row=r, column=ci, value=col)
hdr(ws, r, 3)

for cat, cnt in cat_counts.most_common():
    r += 1
    ws.cell(row=r, column=1, value=cat)
    ws.cell(row=r, column=2, value=cnt)
    c3 = ws.cell(row=r, column=3, value=cnt/len(claims))
    cell_style(ws.cell(row=r, column=1)); cell_style(ws.cell(row=r, column=2)); cell_style(c3, pct=True)

# Portfolio comparison
r += 3
ws.merge_cells(f'A{r}:H{r}')
ws[f'A{r}'] = '▌ PORTFOLIO COMPARISON (Full Auth+Claims Dataset)'
ws[f'A{r}'].font = SUB_FONT

r += 1
for ci, col in enumerate(['Metric','Value'], 1):
    ws.cell(row=r, column=ci, value=col)
hdr(ws, r, 2)

portfolio_rows = [
    ('Total Claims in Portfolio',  len(claim_rows)),
    ('Portfolio Net Claimed (SAR)', f'{all_net:,.2f}'),
    ('Portfolio Approved (SAR)',    f'{all_approved:,.2f}'),
    ('Portfolio Approval Rate',     f'{all_approved/all_net*100:.1f}%' if all_net else 'N/A'),
    ('Batch Approval Rate',         f'{approved_amt/net_claimed*100:.1f}%'),
    ('Batch vs Portfolio Delta',    f'{(approved_amt/net_claimed - all_approved/all_net)*100:+.1f}%' if all_net else 'N/A'),
    ('Most Common Status',          all_statuses.most_common(1)[0][0] if all_statuses else 'N/A'),
]
for ri, (label, val) in enumerate(portfolio_rows, r+1):
    ws.cell(row=ri, column=1, value=label); cell_style(ws.cell(row=ri, column=1))
    ws.cell(row=ri, column=2, value=val);   cell_style(ws.cell(row=ri, column=2))

autowidth(ws)


# ═══════════════════════════════════════════════════════════
# S2 – CLAIMS DATA (full normalized extract)
# ═══════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Claims Data')
ws2.sheet_properties.tabColor = '4472C4'

h2 = ['Bundle ID','Member Name','National ID','Patient ID','Policy No',
      'Claim Type','Claim Status','Service Date','Service Code',
      'Service Name','Rejection Code','Rejection Description','Category',
      'Net Amount (SAR)','Approved Amount (SAR)',
      'PreAuth Required','Preventable','Recommended Action']
for ci, col in enumerate(h2, 1):
    ws2.cell(row=1, column=ci, value=col)
hdr(ws2, 1, len(h2))

for ri, c in enumerate(claims, 2):
    alt = ri % 2 == 0
    vals = [
        c['bundle_id'], c['member_name'], c['national_id'], c['patient_id'],
        c['policy_number'], c['claim_type'], c['claim_status'],
        c['service_date'], c['service_code'], c['service_name'],
        c['rejection_code'], c['rejection_description'], c['category'],
        c['net_amount'], c['approved_amount'],
        c['preauth_required'], c['preventable'], c['action'],
    ]
    for ci, v in enumerate(vals, 1):
        cell = ws2.cell(row=ri, column=ci, value=v)
        cell_style(cell, wrap=(ci in [10,12,18]),
                   money=(ci in [14,15]), alt=alt)
        if ci == 17 and v in PREV_COLOR:
            cell.fill = PREV_COLOR[v]

ws2.auto_filter.ref = f'A1:{get_column_letter(len(h2))}1'
ws2.freeze_panes = 'A2'
autowidth(ws2)


# ═══════════════════════════════════════════════════════════
# S3 – APPEAL TRACKER
# ═══════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Appeal Tracker')
ws3.sheet_properties.tabColor = 'ED7D31'

h3 = ['#','Bundle ID','Member Name','National ID','Rejection Code',
      'Category','Appeal Strategy','Supporting Docs Required',
      'Protocol Section','Est. Recovery (SAR)','Status','Notes']
for ci, col in enumerate(h3, 1):
    ws3.cell(row=1, column=ci, value=col)
hdr(ws3, 1, len(h3))

seen = {}
arow = 2
for c in claims:
    key = (c['bundle_id'], c['rejection_code'])
    if key in seen:
        continue
    seen[key] = True
    est = rej_per_item * c['recovery_rate']
    vals = [
        arow-1, c['bundle_id'], c['member_name'], c['national_id'],
        c['rejection_code'], c['category'],
        c['appeal_strategy'], c['docs_required'], c['protocol_section'],
        round(est, 2), 'PENDING', '',
    ]
    for ci, v in enumerate(vals, 1):
        cell = ws3.cell(row=arow, column=ci, value=v)
        cell_style(cell, wrap=(ci in [7,8,9]), money=(ci==10))
    arow += 1

ws3.auto_filter.ref = f'A1:{get_column_letter(len(h3))}{arow-1}'
ws3.freeze_panes = 'A2'
autowidth(ws3)
print(f"  Appeal Tracker: {arow-2} unique bundle+code combinations")


# ═══════════════════════════════════════════════════════════
# S4 – PRIORITY ACTIONS
# ═══════════════════════════════════════════════════════════
ws4 = wb.create_sheet('Priority Actions')
ws4.sheet_properties.tabColor = '00B050'

ws4.merge_cells('A1:H1')
ws4['A1'] = 'PRIORITY ACTION ITEMS – IMMEDIATE RESPONSE PLAN'
ws4['A1'].font = TITLE_FONT

h4 = ['Priority','Rejection Code','Count','% Impact','Preventable',
      'Requires PreAuth','Immediate Action','Est. Recovery (SAR)','Deadline']
for ci, col in enumerate(h4, 1):
    ws4.cell(row=3, column=ci, value=col)
hdr(ws4, 3, len(h4))

prow = 4
total_recovery = 0
for priority, (code, cnt) in enumerate(rej_counts.most_common(), 1):
    p = ART.get(code, {})
    est = cnt * rej_per_item * p.get('recovery_rate', 0)
    total_recovery += est
    prev = p.get('preventable', 'N/A')
    vals = [
        priority, code, cnt, cnt/len(claims), prev,
        'YES' if p.get('preauth') else 'NO',
        p.get('action','Investigate'),
        round(est, 2), 'Within 30 days'
    ]
    for ci, v in enumerate(vals, 1):
        cell = ws4.cell(row=prow, column=ci, value=v)
        cell_style(cell, wrap=(ci==7), pct=(ci==4), money=(ci==8))
        if ci == 5 and prev in PREV_COLOR:
            cell.fill = PREV_COLOR[prev]
    prow += 1

# Totals row
prow += 1
ws4.cell(row=prow, column=7, value='TOTAL ESTIMATED RECOVERY').font = Font(bold=True)
tc = ws4.cell(row=prow, column=8, value=round(total_recovery, 2))
tc.number_format = MONEY_FMT; tc.font = Font(bold=True, color='00B050')

ws4.cell(row=prow+1, column=7, value='NET EXPOSURE AFTER RECOVERY').font = Font(bold=True)
nc = ws4.cell(row=prow+1, column=8, value=round(rejected_amt - total_recovery, 2))
nc.number_format = MONEY_FMT; nc.font = Font(bold=True, color='FF0000')

autowidth(ws4)


# ═══════════════════════════════════════════════════════════
# S5 – MEMBER ANALYSIS
# ═══════════════════════════════════════════════════════════
ws5 = wb.create_sheet('Member Analysis')
ws5.sheet_properties.tabColor = '7030A0'

h5 = ['National ID','Member Name','Total Rejected Items','Unique Bundles',
      'Primary Code','Codes Affected','Categories','Risk Level']
for ci, col in enumerate(h5, 1):
    ws5.cell(row=1, column=ci, value=col)
hdr(ws5, 1, len(h5))

mbr = defaultdict(lambda: {'names':set(),'bundles':set(),'codes':[],'cats':set()})
for c in claims:
    m = mbr[c['national_id']]
    m['names'].add(c['member_name']); m['bundles'].add(c['bundle_id'])
    m['codes'].append(c['rejection_code']); m['cats'].add(c['category'])

mrow = 2
for nid, m in sorted(mbr.items(), key=lambda x: len(x[1]['codes']), reverse=True):
    cc = Counter(m['codes'])
    primary = cc.most_common(1)[0][0]
    total = len(m['codes'])
    risk = 'HIGH' if total >= 10 else 'MEDIUM' if total >= 5 else 'LOW'
    vals = [nid, ', '.join(sorted(m['names'])), total, len(m['bundles']),
            primary, ', '.join(sorted(set(m['codes']))), ', '.join(sorted(m['cats'])), risk]
    for ci, v in enumerate(vals, 1):
        cell = ws5.cell(row=mrow, column=ci, value=v)
        cell_style(cell, wrap=(ci in [2,6,7]))
        if ci == 8:
            cell.fill = (PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid') if risk=='HIGH'
                        else PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid') if risk=='MEDIUM'
                        else PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'))
    mrow += 1

ws5.auto_filter.ref = f'A1:{get_column_letter(len(h5))}{mrow-1}'
ws5.freeze_panes = 'A2'
autowidth(ws5)


# ═══════════════════════════════════════════════════════════
# S6 – PROVIDER / PORTFOLIO COMPARISON
# ═══════════════════════════════════════════════════════════
ws6 = wb.create_sheet('Portfolio Comparison')
ws6.sheet_properties.tabColor = '002060'

ws6.merge_cells('A1:G1')
ws6['A1'] = 'PORTFOLIO COMPARISON – Batch vs Full Dataset (Al-Rajhi Claims Feb 2026)'
ws6['A1'].font = TITLE_FONT

r6 = 3
ws6.merge_cells(f'A{r6}:G{r6}')
ws6[f'A{r6}'] = '▌ CLAIM TYPE APPROVAL RATES (Full Portfolio)'
ws6[f'A{r6}'].font = SUB_FONT

r6 += 1
for ci, col in enumerate(['Claim Type','Total Claims','Net Amount (SAR)',
                           'Approved Amount (SAR)','Approval Rate',
                           'Avg Claim Value (SAR)'], 1):
    ws6.cell(row=r6, column=ci, value=col)
hdr(ws6, r6, 6)

# Group portfolio by claim type
ct_data = defaultdict(lambda: {'cnt':0,'net':0.0,'appr':0.0})
for d in claim_map.values():
    ct = d.get('Claim Type') or 'Unknown'
    ct_data[ct]['cnt'] += 1
    ct_data[ct]['net'] += d.get('Net Amount') or 0
    ct_data[ct]['appr'] += d.get('Approved Amount') or 0

for ct, dt in sorted(ct_data.items(), key=lambda x: x[1]['net'], reverse=True):
    r6 += 1
    rate = dt['appr']/dt['net'] if dt['net'] else 0
    avg  = dt['net']/dt['cnt'] if dt['cnt'] else 0
    vals = [ct, dt['cnt'], round(dt['net'],2), round(dt['appr'],2), rate, round(avg,2)]
    for ci, v in enumerate(vals, 1):
        cell = ws6.cell(row=r6, column=ci, value=v)
        cell_style(cell, money=(ci in [3,4,6]), pct=(ci==5))
        if ci == 5:  # color code approval rate
            if rate >= 0.8:   cell.fill = PREV_COLOR['YES']
            elif rate >= 0.5: cell.fill = PREV_COLOR['PARTIALLY']
            else:             cell.fill = PREV_COLOR['NO']

# Status breakdown
r6 += 3
ws6.merge_cells(f'A{r6}:G{r6}')
ws6[f'A{r6}'] = '▌ CLAIM STATUS DISTRIBUTION (Full Portfolio)'
ws6[f'A{r6}'].font = SUB_FONT

r6 += 1
for ci, col in enumerate(['Status','Count','% of Portfolio'], 1):
    ws6.cell(row=r6, column=ci, value=col)
hdr(ws6, r6, 3)

total_portfolio = len(claim_rows)
for status, sc in all_statuses.most_common():
    r6 += 1
    ws6.cell(row=r6, column=1, value=status)
    ws6.cell(row=r6, column=2, value=sc)
    c3 = ws6.cell(row=r6, column=3, value=sc/total_portfolio if total_portfolio else 0)
    cell_style(ws6.cell(row=r6,column=1)); cell_style(ws6.cell(row=r6,column=2)); cell_style(c3, pct=True)

# Rejected bundles found in portfolio
r6 += 3
ws6.merge_cells(f'A{r6}:G{r6}')
ws6[f'A{r6}'] = '▌ REJECTED BUNDLES CROSS-REFERENCE'
ws6[f'A{r6}'].font = SUB_FONT

r6 += 1
matched = [(bid, claim_map[bid]) for bid in rejected_bundles if bid in claim_map]
for ci, col in enumerate(['Bundle ID','Claim Status','Net Amount (SAR)',
                           'Approved Amount (SAR)','Claim Type'], 1):
    ws6.cell(row=r6, column=ci, value=col)
hdr(ws6, r6, 5)

for bid, cr in sorted(matched, key=lambda x: x[1].get('Net Amount') or 0, reverse=True):
    r6 += 1
    vals = [bid, cr.get('Status'), cr.get('Net Amount'), cr.get('Approved Amount'), cr.get('Claim Type')]
    for ci, v in enumerate(vals, 1):
        cell = ws6.cell(row=r6, column=ci, value=v)
        cell_style(cell, money=(ci in [3,4]))

autowidth(ws6)


# ═══════════════════════════════════════════════════════════
# S7 – ART PROTOCOL REFERENCE
# ═══════════════════════════════════════════════════════════
ws7 = wb.create_sheet('ART Protocol Reference')
ws7.sheet_properties.tabColor = 'BF8F00'

h7 = ['Code','Description','Protocol Section','Requires PreAuth',
      'Preventable','Recovery Rate','Supporting Docs Required',
      'Appeal Strategy','Typical Services']
for ci, col in enumerate(h7, 1):
    ws7.cell(row=1, column=ci, value=col)
hdr(ws7, 1, len(h7))

for prow, (code, p) in enumerate(ART.items(), 2):
    vals = [
        code, p['desc'], p['section'],
        'YES' if p['preauth'] else 'NO',
        p['preventable'], p['recovery_rate'],
        p['docs'], p['strategy'],
        ', '.join(p.get('services',[])),
    ]
    for ci, v in enumerate(vals, 1):
        cell = ws7.cell(row=prow, column=ci, value=v)
        cell_style(cell, wrap=(ci in [2,7,8,9]), pct=(ci==6))
        if ci == 5 and v in PREV_COLOR:
            cell.fill = PREV_COLOR[v]

autowidth(ws7)


# ═══════════════════════════════════════════════════════════
# S8 – TREND & PATTERNS
# ═══════════════════════════════════════════════════════════
ws8 = wb.create_sheet('Trend & Patterns')
ws8.sheet_properties.tabColor = 'FF0000'

ws8.merge_cells('A1:F1')
ws8['A1'] = 'TREND & PATTERN DISCOVERY'
ws8['A1'].font = TITLE_FONT

# Cross-tab: code × category
r8 = 3
ws8.merge_cells(f'A{r8}:F{r8}')
ws8[f'A{r8}'] = '▌ REJECTION CODE × CATEGORY MATRIX'
ws8[f'A{r8}'].font = SUB_FONT

r8 += 1
cats_sorted = sorted(set(c['category'] for c in claims))
ws8.cell(row=r8, column=1, value='Code \\ Category')
for ci, cat in enumerate(cats_sorted, 2):
    ws8.cell(row=r8, column=ci, value=cat)
ws8.cell(row=r8, column=len(cats_sorted)+2, value='TOTAL')
hdr(ws8, r8, len(cats_sorted)+2)

cross = defaultdict(lambda: defaultdict(int))
for c in claims:
    cross[c['rejection_code']][c['category']] += 1

for code in sorted(cross):
    r8 += 1
    ws8.cell(row=r8, column=1, value=code)
    for ci, cat in enumerate(cats_sorted, 2):
        v = cross[code].get(cat, 0)
        if v:
            ws8.cell(row=r8, column=ci, value=v)
    ws8.cell(row=r8, column=len(cats_sorted)+2, value=sum(cross[code].values()))
    for ci in range(1, len(cats_sorted)+3):
        cell_style(ws8.cell(row=r8, column=ci))

# Date distribution
r8 += 3
ws8.merge_cells(f'A{r8}:F{r8}')
ws8[f'A{r8}'] = '▌ SERVICE DATE DISTRIBUTION'
ws8[f'A{r8}'].font = SUB_FONT

r8 += 1
for ci, col in enumerate(['Service Date','Rejections','Unique Members','% of Total'], 1):
    ws8.cell(row=r8, column=ci, value=col)
hdr(ws8, r8, 4)

date_data = defaultdict(lambda: {'count':0,'members':set()})
for c in claims:
    date_data[c['service_date']]['count'] += 1
    date_data[c['service_date']]['members'].add(c['national_id'])

for d in sorted(date_data):
    r8 += 1
    dd = date_data[d]
    vals = [d, dd['count'], len(dd['members']), dd['count']/len(claims)]
    for ci, v in enumerate(vals, 1):
        cell = ws8.cell(row=r8, column=ci, value=v)
        cell_style(cell, pct=(ci==4))

# Financial exposure summary
r8 += 3
ws8.merge_cells(f'A{r8}:F{r8}')
ws8[f'A{r8}'] = '▌ FINANCIAL EXPOSURE ANALYSIS'
ws8[f'A{r8}'].font = SUB_FONT

r8 += 1
for ci, col in enumerate(['Metric','SAR'], 1):
    ws8.cell(row=r8, column=ci, value=col)
hdr(ws8, r8, 2)

exposure = [
    ('Net Claimed (Batch)',                  net_claimed),
    ('Approved Amount (Batch)',              approved_amt),
    ('Rejected Amount (Batch)',              rejected_amt),
    ('VAT on Approved',                      vat_approved),
    ('Gross Variance',                       gross_var),
    ('Avg Rejection per Line Item',          rej_per_item),
    ('Total Portfolio Net (Feb dataset)',     all_net),
    ('Total Portfolio Approved',             all_approved),
    ('Portfolio Rejection (estimated)',      all_net - all_approved),
    ('Estimated Recovery via Appeals',       total_recovery),
    ('Net Exposure After Recovery',          rejected_amt - total_recovery),
]
for rr, (metric, val) in enumerate(exposure, r8+1):
    ws8.cell(row=rr, column=1, value=metric)
    c = ws8.cell(row=rr, column=2, value=round(val,2) if val else 0)
    cell_style(ws8.cell(row=rr, column=1))
    cell_style(c, money=True)

autowidth(ws8)


# ═══════════════════════════════════════════════════════════
# S9 – SUBMISSION QUEUE
# ═══════════════════════════════════════════════════════════
ws9 = wb.create_sheet('Submission Queue')
ws9.sheet_properties.tabColor = 'FF0000'

ws9.merge_cells('A1:K1')
ws9['A1'] = 'NPHIES APPEAL SUBMISSION QUEUE – Ready for CommunicationRequest'
ws9['A1'].font = TITLE_FONT

h9 = ['#','Bundle ID','Transaction ID','Member Name','National ID',
      'Rejection Code','Appeal Strategy','Readiness',
      'Est. Recovery (SAR)','Supporting Docs','Status']
for ci, col in enumerate(h9, 1):
    ws9.cell(row=2, column=ci, value=col)
hdr(ws9, 2, len(h9))

seen9 = {}
qrow = 3
for c in claims:
    key = c['bundle_id']
    if key in seen9:
        continue
    seen9[key] = True
    code = c['rejection_code']
    p = ART.get(code, {})
    prev = p.get('preventable','N/A')
    readiness = ('READY_AUTO_APPEAL'   if prev == 'YES'
                 else 'PARTIAL_AUTO_APPEAL' if prev == 'PARTIALLY'
                 else 'MANUAL_REVIEW')
    est = rej_per_item * p.get('recovery_rate', 0)
    vals = [
        qrow-2, c['bundle_id'], c['transaction_id'], c['member_name'],
        c['national_id'], code, p.get('strategy','Manual review'),
        readiness, round(est,2), p.get('docs','Clinical documentation'), 'PENDING'
    ]
    for ci, v in enumerate(vals, 1):
        cell = ws9.cell(row=qrow, column=ci, value=v)
        cell_style(cell, wrap=(ci in [7,10]), money=(ci==9))
        # Color by readiness
        if ci == 8:
            if readiness == 'READY_AUTO_APPEAL':
                cell.fill = PREV_COLOR['YES']
            elif readiness == 'PARTIAL_AUTO_APPEAL':
                cell.fill = PREV_COLOR['PARTIALLY']
            else:
                cell.fill = PREV_COLOR['NO']
    qrow += 1

ws9.auto_filter.ref = f'A2:{get_column_letter(len(h9))}{qrow-1}'
ws9.freeze_panes = 'A3'
autowidth(ws9)

ready_count    = sum(1 for c in claims if ART.get(c['rejection_code'],{}).get('preventable')=='YES')
partial_count  = sum(1 for c in claims if ART.get(c['rejection_code'],{}).get('preventable')=='PARTIALLY')
manual_count   = sum(1 for c in claims if ART.get(c['rejection_code'],{}).get('preventable') not in ('YES','PARTIALLY'))

print(f"  Submission Queue: {len(seen9)} unique bundles | "
      f"Auto={ready_count} Partial={partial_count} Manual={manual_count}")


# ─── Save workbook ────────────────────────────────────────────────────────────
wb.save(OUT_XLS)
print(f"\n✅ Workbook saved: {OUT_XLS}")
print(f"   Sheets: {wb.sheetnames}")


# ─── Export appeal-prep CSVs ──────────────────────────────────────────────────
import csv

# claim_appeal_summary.csv (one per bundle)
summary_file = f"{PREP_DIR}/claim_appeal_summary.csv"
with open(summary_file, 'w', newline='', encoding='utf-8') as f:
    wr = csv.writer(f)
    wr.writerow(['BundleID','TransactionID','MemberName','NationalID','PolicyNo',
                 'ClaimType','ClaimStatus','RejectionCode','Category',
                 'AppealStrategy','Readiness','EstRecovery'])
    seen_s = set()
    for c in claims:
        if c['bundle_id'] in seen_s: continue
        seen_s.add(c['bundle_id'])
        p = ART.get(c['rejection_code'], {})
        prev = p.get('preventable','N/A')
        readiness = ('READY_AUTO_APPEAL' if prev=='YES'
                     else 'PARTIAL_AUTO_APPEAL' if prev=='PARTIALLY'
                     else 'MANUAL_REVIEW')
        est = rej_per_item * p.get('recovery_rate', 0)
        wr.writerow([c['bundle_id'], c['transaction_id'], c['member_name'],
                     c['national_id'], c['policy_number'], c['claim_type'],
                     c['claim_status'], c['rejection_code'], c['category'],
                     p.get('strategy','Manual review'), readiness, round(est,2)])

print(f"   Appeal Summary CSV: {summary_file} ({len(seen_s)} bundles)")

# line_appeal_detail.csv (one per rejected line)
detail_file = f"{PREP_DIR}/line_appeal_detail.csv"
with open(detail_file, 'w', newline='', encoding='utf-8') as f:
    wr = csv.writer(f)
    wr.writerow(['BundleID','MemberName','NationalID','ServiceDate','ServiceCode',
                 'ServiceName','RejectionCode','RejectionDescription','Category',
                 'PreAuthRequired','Preventable','DocsRequired','AppealStrategy'])
    for c in claims:
        wr.writerow([c['bundle_id'], c['member_name'], c['national_id'],
                     c['service_date'], c['service_code'], c['service_name'],
                     c['rejection_code'], c['rejection_description'], c['category'],
                     c['preauth_required'], c['preventable'],
                     c['docs_required'], c['appeal_strategy']])

print(f"   Line Detail CSV: {detail_file} ({len(claims)} lines)")

# denial_category_summary.csv
cat_file = f"{PREP_DIR}/denial_category_summary.csv"
with open(cat_file, 'w', newline='', encoding='utf-8') as f:
    wr = csv.writer(f)
    wr.writerow(['Category','Count','PctOfTotal','EstRecovery'])
    for cat, cnt in cat_counts.most_common():
        est = sum(rej_per_item * ART.get(c['rejection_code'],{}).get('recovery_rate',0)
                  for c in claims if c['category']==cat)
        wr.writerow([cat, cnt, f'{cnt/len(claims)*100:.1f}%', round(est,2)])

print(f"   Category Summary CSV: {cat_file}")

# appeal_summary.json (for FHIR bundle generation)
appeal_json = f"{PREP_DIR}/appeal_summary.json"
payer_info = {
    'name': 'Al-Rajhi Company for Cooperative Insurance',
    'license': '7001593321',
    'system': 'http://nphies.sa/identifier/payer',
}
provider_info = {
    'name': 'Al-Hayat National Hospital',
    'license': '10000000000988',
    'system': 'http://nphies.sa/identifier/chi-license',
}
with open(appeal_json, 'w') as f:
    json.dump({
        'batch_no': header['batch_no'],
        'period': f"{header['from_date']} - {header['to_date']}",
        'generated': datetime.now().isoformat(),
        'payer': payer_info,
        'provider': provider_info,
        'financial_summary': {
            'net_claimed': net_claimed,
            'rejected_amount': rejected_amt,
            'approved_amount': approved_amt,
            'estimated_recovery': round(total_recovery, 2),
            'net_exposure': round(rejected_amt - total_recovery, 2),
        },
        'stats': {
            'total_rejected_lines': len(claims),
            'unique_bundles': len(rejected_bundles),
            'unique_members': len(mbr),
            'ready_auto': sum(1 for c in claims if ART.get(c['rejection_code'],{}).get('preventable')=='YES'),
            'partial_auto': sum(1 for c in claims if ART.get(c['rejection_code'],{}).get('preventable')=='PARTIALLY'),
            'manual_review': sum(1 for c in claims if ART.get(c['rejection_code'],{}).get('preventable') not in ('YES','PARTIALLY')),
        },
        'rejection_codes': {code: {'count': cnt, 'description': ART.get(code,{}).get('desc',''),
                                    'preventable': ART.get(code,{}).get('preventable',''),
                                    'appeal_strategy': ART.get(code,{}).get('strategy','')}
                             for code, cnt in rej_counts.most_common()},
        'claims': [{
            'bundle_id': c['bundle_id'], 'transaction_id': c['transaction_id'],
            'member_name': c['member_name'], 'national_id': c['national_id'],
            'policy_number': c['policy_number'], 'claim_type': c['claim_type'],
            'rejection_code': c['rejection_code'], 'category': c['category'],
            'service_date': c['service_date'], 'service_name': c['service_name'],
            'service_code': c['service_code'], 'net_amount': c['net_amount'],
            'appeal_strategy': c['appeal_strategy'], 'docs_required': c['docs_required'],
            'preventable': c['preventable'],
        } for c in claims],
    }, f, indent=2, ensure_ascii=False)

print(f"   Appeal JSON: {appeal_json}")
print(f"\n{'='*60}")
print(f"  BATCH: {header['batch_no']}")
print(f"  Net Claimed:       SAR {net_claimed:>12,.2f}")
print(f"  Rejected:          SAR {rejected_amt:>12,.2f}  ({rejected_amt/net_claimed*100:.1f}%)")
print(f"  Approved:          SAR {approved_amt:>12,.2f}  ({approved_amt/net_claimed*100:.1f}%)")
print(f"  Est. Recovery:     SAR {total_recovery:>12,.2f}  ({total_recovery/rejected_amt*100:.1f}%)")
print(f"  Net Exposure:      SAR {rejected_amt-total_recovery:>12,.2f}")
print(f"  Rejected Items:    {len(claims)}")
print(f"  Unique Bundles:    {len(rejected_bundles)}")
print(f"  Ready Auto:        {ready_count}")
print(f"  Partial Auto:      {partial_count}")
print(f"  Manual Review:     {manual_count}")
print(f"{'='*60}")
