#!/usr/bin/env python3
"""
Rebuild NPHIES FHIR CommunicationRequest appeal bundles using REAL identifiers
from the Claims Excel export (Claim_9f332ffe...xlsx).

Key fix: Use Transaction Identifier (NPHIES claim number) in the about[] reference,
not the UUID bundle_id. Use proper integer Patient/Provider/Receiver license values.
"""
import json, csv, os, re
from datetime import datetime, timezone
from collections import defaultdict
import openpyxl

# ─── Paths ───────────────────────────────────────────────────────────────────
BASE       = "/Users/fadil369/abha-nphies-session-deliverables-1"
DL         = "/Users/fadil369/Downloads/rajhi-riyad-feb"
CLAIM_XLS  = f"{DL}/Claim_9f332ffe-7145-49ee-aa78-5dc56d4cdfc1_part1_2026-02-07_2026-03-10.xlsx"
PREP_JSON  = f"{BASE}/outputs/rajhi-appeal-prep/appeal_summary.json"
OUT_DIR    = f"{BASE}/outputs/rajhi-appeal-execution"
BUNDLE_DIR = f"{OUT_DIR}/bundles"
os.makedirs(BUNDLE_DIR, exist_ok=True)

# ─── ART protocol action text ─────────────────────────────────────────────────
ACTION_TEXT = {
    'BE-1-4': (
        "PREAUTHORIZATION APPEAL: The service was provided without prior authorization "
        "due to clinical urgency / administrative oversight. We hereby submit a retroactive "
        "preauthorization request with full clinical justification as required under ART PreAuth "
        "Protocol – General Requirements. All supporting clinical documentation is attached."
    ),
    'MN-1-1': (
        "CLINICAL JUSTIFICATION APPEAL: The service is clinically justified as evidenced by the "
        "attached medical records, diagnosis codes, and clinical pathway documentation. The service "
        "aligns with accepted Clinical Practice Guidelines (CPG) and was medically necessary."
    ),
    'CV-1-3': (
        "COVERAGE VERIFICATION APPEAL: We are appealing the coverage determination and requesting "
        "re-adjudication based on the attached policy benefit schedule. If the primary diagnosis is "
        "excluded, we request consideration of the secondary covered diagnosis as documented in "
        "the clinical notes."
    ),
    'BE-1-3': (
        "CONTRACTUAL COMPLIANCE APPEAL: The service codes have been reviewed against the "
        "contractual agreement. Corrected service codes are provided per the agreed tariff "
        "schedule. We request re-adjudication under the correct code mapping."
    ),
    'SE-1-6': (
        "SUPPORTING EVIDENCE APPEAL: The missing investigation results (laboratory/radiology) "
        "are attached to this communication as supporting evidence. These results confirm the "
        "medical necessity and clinical justification for the service rendered."
    ),
    'AD-1-4': (
        "DIAGNOSIS-SERVICE ALIGNMENT APPEAL: The diagnosis codes have been reviewed and "
        "corrected to accurately reflect the clinical picture. The corrected ICD-10 codes are "
        "consistent with the service provided. We request re-adjudication with the updated coding."
    ),
    'CV-1-9': (
        "COVERAGE APPEAL: Member eligibility and benefit coverage have been verified. "
        "We request re-adjudication under the verified coverage confirmation attached."
    ),
    'AD-3-7': (
        "ADDITIONAL INFORMATION APPEAL: The additional documentation requested for adjudication "
        "is attached herewith – clinical records, reports, and supporting materials required "
        "for complete adjudication."
    ),
    'AD-2-4': (
        "DUPLICATE REVIEW APPEAL: We have verified that the original claim has not been settled. "
        "This claim represents a unique service encounter and should be adjudicated independently."
    ),
}

ART_DOCS = {
    'BE-1-4': 'Medical report, clinical notes, diagnosis confirmation, treatment plan',
    'MN-1-1': 'Lab results, clinical notes, CPG reference, medical necessity letter',
    'CV-1-3': 'Policy benefit schedule, clinical documentation for alternate diagnosis',
    'BE-1-3': 'Correct NPHIES service code mapping, contract reference',
    'SE-1-6': 'X-ray images, lab reports, dental charting, clinical photos',
    'AD-1-4': 'Updated clinical notes with correct ICD-10 codes, treatment rationale',
    'CV-1-9': 'Eligibility verification, benefit schedule confirmation',
    'AD-3-7': 'Medical reports, clinical notes, additional documentation as requested',
    'AD-2-4': 'Original claim reference, settlement reconciliation report',
}

ART_PREVENTABLE = {
    'BE-1-4': 'YES', 'MN-1-1': 'PARTIALLY', 'CV-1-3': 'NO',
    'BE-1-3': 'YES', 'SE-1-6': 'YES',       'AD-1-4': 'PARTIALLY',
    'CV-1-9': 'PARTIALLY', 'AD-3-7': 'YES', 'AD-2-4': 'YES',
}


# ─── Load Claims Excel (real NPHIES identifiers) ───────────────────────────────
print("Loading Claims Excel (real NPHIES identifiers)...")
wb = openpyxl.load_workbook(CLAIM_XLS, read_only=True)
ws = wb.active
headers = list(ws.iter_rows(max_row=1, values_only=True))[0]
rows = list(ws.iter_rows(min_row=2, values_only=True))
claim_data = {str(row[headers.index('Bundle ID')]): dict(zip(headers, row)) for row in rows if row[0]}
wb.close()
print(f"  {len(claim_data)} bundle records loaded")

# ─── Load appeal prep JSON ─────────────────────────────────────────────────────
with open(PREP_JSON, encoding='utf-8') as f:
    prep = json.load(f)

payer    = prep['payer']
provider = prep['provider']
claims   = prep['claims']

# Group by bundle_id
bundle_claims = defaultdict(list)
for c in claims:
    bundle_claims[c['bundle_id']].append(c)

print(f"  Batch bundles: {len(bundle_claims)}")

# ─── Rebuild bundles ──────────────────────────────────────────────────────────
now_ts  = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
results = []

for bundle_id, bundle_cls in bundle_claims.items():
    first = bundle_cls[0]

    # Pull REAL identifiers from Claims Excel
    cr = claim_data.get(bundle_id, {})

    # Clean up float → int → str conversions from Excel
    def clean_id(val):
        if val is None:
            return ''
        s = str(val)
        return s[:-2] if s.endswith('.0') else s

    transaction_id    = clean_id(cr.get('Transaction Identifier'))
    patient_id        = clean_id(cr.get('Patient Identifier'))
    provider_license  = clean_id(cr.get('Provider License'))  or provider['license']
    receiver_license  = clean_id(cr.get('Receiver License'))  or payer['license']
    claim_type        = cr.get('Claim Type')        or 'professional'
    claim_sub_type    = cr.get('Claim Sub Type')    or 'op'
    policy_number     = clean_id(cr.get('Policy Number'))
    policy_holder     = cr.get('Policy Holder Name') or ''
    encounter_class   = cr.get('Encounter Class')   or ''
    net_amount        = cr.get('Net Amount')        or 0.0
    insurer_license   = clean_id(cr.get('Insurer License')) or payer['license']
    insurer_name      = cr.get('Insurer Name') or payer['name']
    receiver_name     = cr.get('Receiver Name') or payer['name']
    provider_name     = (cr.get('Provider Name') or provider['name']).strip().rstrip(',')

    # Determine readiness
    primary_codes = list(dict.fromkeys(c['rejection_code'] for c in bundle_cls))
    primary_code  = primary_codes[0]
    preventable   = ART_PREVENTABLE.get(primary_code, 'N/A')
    readiness = (
        'READY_AUTO_APPEAL'   if preventable == 'YES'
        else 'PARTIAL_AUTO_APPEAL' if preventable == 'PARTIALLY'
        else 'MANUAL_REVIEW'
    )

    # Group by rejection code
    by_code = defaultdict(list)
    for c in bundle_cls:
        by_code[c['rejection_code']].append(c)

    # Build payload text
    lines = [
        f"RE-ADJUDICATION REQUEST",
        f"Batch: {prep['batch_no']}  |  Period: {prep['period']}",
        f"",
        f"CLAIM DETAILS:",
        f"  Bundle ID:          {bundle_id}",
        f"  Transaction ID:     {transaction_id}",
        f"  Patient NID:        {patient_id}",
        f"  Claim Type:         {claim_type} / {claim_sub_type}",
        f"  Encounter Class:    {encounter_class}",
        f"  Policy Number:      {policy_number}",
        f"  Policy Holder:      {policy_holder}",
        f"  Net Amount:         SAR {net_amount:,.2f}" if isinstance(net_amount, float) else f"  Net Amount:         SAR {net_amount}",
        f"",
        f"REJECTION DETAILS:",
    ]

    for code, code_claims in sorted(by_code.items()):
        lines.append(f"  [{code}] – {len(code_claims)} service(s) rejected:")
        for cc in code_claims:
            lines.append(f"    • {cc['service_date']} | {cc['service_code']} – {cc['service_name']}")
        lines.append(f"")
        lines.append(f"  APPEAL JUSTIFICATION ({code}):")
        lines.append(f"  {ACTION_TEXT.get(code, 'Re-adjudication requested with supporting documentation.')}")
        lines.append(f"  Supporting Documents: {ART_DOCS.get(code, 'Clinical documentation')}")
        lines.append(f"")

    lines += [
        "---",
        f"Provider: {provider_name}",
        f"Provider License: {provider_license}",
        f"Payer: {receiver_name}",
        f"Payer License: {receiver_license}",
    ]

    # Build FHIR CommunicationRequest with REAL identifiers
    bundle = {
        "facility_id": 1,
        "fhir_payload": {
            "resourceType": "CommunicationRequest",
            "id": f"appeal-{transaction_id}-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}",
            "status": "active",
            "category": [
                {
                    "coding": [
                        {
                            "system": "http://nphies.sa/terminology/CodeSystem/communication-category",
                            "code": "re-adjudication",
                            "display": "Re-Adjudication Request"
                        }
                    ]
                }
            ],
            "priority": "routine",
            "subject": {
                "identifier": {
                    "system": "http://nphies.sa/identifier/patient",
                    "value": patient_id
                },
                "display": first['member_name']
            },
            "about": [
                {
                    "type": "Claim",
                    "identifier": {
                        "system": "http://nphies.sa/identifier/claim",
                        "value": bundle_id
                    }
                },
                {
                    "type": "Claim",
                    "identifier": {
                        "system": "http://nphies.sa/identifier/transaction",
                        "value": transaction_id
                    }
                }
            ],
            "sender": {
                "type": "Organization",
                "identifier": {
                    "system": "http://nphies.sa/identifier/chi-license",
                    "value": provider_license
                },
                "display": provider_name
            },
            "recipient": [
                {
                    "type": "Organization",
                    "identifier": {
                        "system": "http://nphies.sa/identifier/payer",
                        "value": receiver_license
                    },
                    "display": receiver_name
                }
            ],
            "reasonCode": [
                {
                    "coding": [
                        {
                            "system": "http://nphies.sa/terminology/CodeSystem/rejection-reason",
                            "code": code
                        }
                    ]
                }
                for code in primary_codes[:3]  # up to 3 rejection codes
            ],
            "payload": [
                {
                    "contentString": "\n".join(lines)
                }
            ],
            "authoredOn": now_ts
        },
        "signature": "",
        "resource_type": "CommunicationRequest"
    }

    # Write bundle
    safe_id  = re.sub(r'[^a-zA-Z0-9_-]', '_', bundle_id)
    filename = f"appeal_{safe_id}_{readiness}.json"
    with open(f"{BUNDLE_DIR}/{filename}", 'w', encoding='utf-8') as f:
        json.dump(bundle, f, indent=2, ensure_ascii=False)

    results.append({
        'bundle_id':        bundle_id,
        'transaction_id':   transaction_id,
        'patient_id':       patient_id,
        'provider_license': provider_license,
        'receiver_license': receiver_license,
        'policy_number':    policy_number,
        'policy_holder':    policy_holder,
        'claim_type':       claim_type,
        'rejection_codes':  ', '.join(primary_codes),
        'lines_rejected':   len(bundle_cls),
        'readiness':        readiness,
        'net_amount':       net_amount,
        'filename':         filename,
        'status':           'GENERATED',
    })

print(f"\nRebuilt {len(results)} bundles with real NPHIES identifiers")

# Readiness summary
from collections import Counter
by_r = Counter(r['readiness'] for r in results)
print(f"  READY_AUTO_APPEAL:   {by_r.get('READY_AUTO_APPEAL', 0)}")
print(f"  PARTIAL_AUTO_APPEAL: {by_r.get('PARTIAL_AUTO_APPEAL', 0)}")
print(f"  MANUAL_REVIEW:       {by_r.get('MANUAL_REVIEW', 0)}")

# Missing transaction IDs
missing = [r for r in results if not r['transaction_id']]
if missing:
    print(f"\n⚠️  {len(missing)} bundles missing Transaction ID:")
    for r in missing:
        print(f"  {r['bundle_id']}")

# Update execution report
report_path = f"{OUT_DIR}/execution_report.json"
with open(report_path, encoding='utf-8') as f:
    report = json.load(f)
report['bundles'] = results
report['rebuilt_at'] = datetime.now(timezone.utc).isoformat()
report['identifier_source'] = 'Claims Excel export (real NPHIES transaction IDs)'
report['totals']['READY_AUTO_APPEAL']   = by_r.get('READY_AUTO_APPEAL', 0)
report['totals']['PARTIAL_AUTO_APPEAL'] = by_r.get('PARTIAL_AUTO_APPEAL', 0)
report['totals']['MANUAL_REVIEW']       = by_r.get('MANUAL_REVIEW', 0)
with open(report_path, 'w', encoding='utf-8') as f:
    json.dump(report, f, indent=2, ensure_ascii=False)

# Update CSV
with open(f"{OUT_DIR}/execution_results.csv", 'w', newline='', encoding='utf-8') as f:
    wr = csv.DictWriter(f, fieldnames=results[0].keys())
    wr.writeheader()
    wr.writerows(results)

# Sample verification
print("\nSample bundle verification:")
r0 = results[0]
print(f"  Bundle ID:        {r0['bundle_id']}")
print(f"  Transaction ID:   {r0['transaction_id']}")
print(f"  Patient ID:       {r0['patient_id']}")
print(f"  Provider License: {r0['provider_license']}")
print(f"  Receiver License: {r0['receiver_license']}")
print(f"  Policy Number:    {r0['policy_number']}")
print(f"  Readiness:        {r0['readiness']}")
print(f"\n✅ Done. Bundles: {OUT_DIR}/bundles/")
